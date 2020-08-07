#!/usr/bin/env python
# encoding: utf-8
'''
@author: yanghong
@file: readexceltools.py
@time: 2020/5/9 16:14
@desc:Excel数据读取
'''
from openpyxl import load_workbook


class Excel_read:
    def __init__(self, filepath, sheetname):
        self.wk = load_workbook(filepath)
        # self.sheet=self.wk.get_sheet_by_name(sheetname)
        self.sheet = self.wk[sheetname]  # 新方法
        self.maxrow = self.sheet.max_row  # 最大行数
        self.maxcolumn = self.sheet.max_column  # 最大列数

    def get_dict_data(self):
        if self.maxrow <= 1:
            raise ValueError('Excel文件内容总行数≤1')
        else:
            dic_list = []
            for i in range(2, self.maxrow + 1):
                d = {}
                d['rowNum'] = i
                for j in range(1, self.maxcolumn + 1):
                    key = self.sheet.cell(1, j).value
                    if self.sheet.cell(i, j).value == None:
                        d[key] = ''
                    else:
                        d[key] = self.sheet.cell(i, j).value.replace(' ', '').replace('\n', '')
                dic_list.append(d)
            return (dic_list)


if __name__ == '__main__':
    import os

    datapath = os.path.join(os.path.dirname(__file__), 'ExcelFile')
    filepath = os.path.join(datapath, 'debug_api.xlsx')
    a = Excel_read(filepath, 'Sheet1')
    for i in a.get_dict_data():
        print(i)
